import importlib
import subprocess
import time
import platform
import os
import sys
import random
from pathlib import Path
from datetime import datetime




# Get the operating system information
operating_system = platform.system()

# Print the operating system
print(f"\nThe operating system is: {operating_system}")




################################################################################################################################


# List of required modules
required_modules = ['selenium', 'webdriver_manager', 'pyperclip', 'undetected_chromedriver', 'requests', 'openpyxl', 'iso-639']

print("\nrequired_modules : ",required_modules)

time.sleep(3)




for required_module in required_modules:
    
    try:
       required_module_ = required_module.replace('-', '')
       print("\n_module_ : ",required_module_)
       importlib.import_module(required_module_)
       print(f"\n{required_module} is already installed.")
    except:
        if 'darwin' in str(operating_system).strip().lower() or 'mac' in str(operating_system).strip().lower():
            try:
                os.system(f"pip3 install {required_module}")
                importlib.import_module(required_module_)
            except:
                print(f"\nFaced issues with {required_module}")
                time.sleep(600)
                
        elif 'linux' in str(operating_system).strip().lower():
            # Linux specific installation
            try:
                # Use appropriate package manager for Linux (e.g., apt for Debian/Ubuntu based)
                print(f"\nsudo apt-get install {required_module}")
                os.system(f"sudo apt-get install {required_module}")
                importlib.import_module(required_module_)
                print(f"\n{required_module} installed successfully.")
            except:
                print(f"\nFaced issues with installing {required_module} on Linux")
                time.sleep(600)
                
        else:
            try:
                os.system(f"pip install {required_module}")
                import selenium
                print(f"\n{required_module} installed successfully.")
            except:
                print(f"\nFaced issues with {required_module}")
                time.sleep(600)
            




    
################################################################################################################################

import openpyxl
import requests


import json
import webdriver_manager
import selenium
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import undetected_chromedriver as uc
from selenium.webdriver import ActionChains
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException, \
        WebDriverException, InvalidSessionIdException
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
import selenium.webdriver.common.keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.alert import Alert
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC, expected_conditions
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.keys import Keys

import urllib.request
from urllib.request import urlopen
import ssl
import json
ssl._create_default_https_context = ssl._create_unverified_context


cmd_ctrl = Keys.COMMAND if sys.platform == 'darwin' else Keys.CONTROL


import pyperclip
import multiprocessing
from multiprocessing import Process, Manager
import functools
import iso639
from iso639 import languages



################################################################################################################################







            
            
            
################################################################################################################################


def quora_post(the_row, the_link_to_add, all_languages,wb_Settings,ws_Settings, clipboard):
    
    
    # Reading from the text file and creating a new dictionary
    loaded_language_codes = {}
    with open('google_translate_codes.txt', 'r', encoding='utf-8') as file:
        lines = file.readlines()
        for line in lines:
            # Splitting each line into language (lowercase) and code
            language, code = line.strip().split(':')
            language =language.lower().strip()
            code = code.lower().strip()
            loaded_language_codes[language] = code
            
            

    # Function to get language code based on lowercase language name
    def get_language_code(language_name):
        lower_language_name = language_name.lower()
        if lower_language_name in loaded_language_codes:
            return loaded_language_codes[lower_language_name]
        else:
            return None
    
    
    path = ''
    the_folder_name = ''
    
    try:
        print("\nThe Row No. : ", the_row)
        
        the_folder_name = ws_Settings.cell(row= the_row,column= 2).value
        
        print("\nthe_folder_name : ",the_folder_name)
        
        if the_folder_name == '' or the_folder_name == None:
            print("\nThe folder name's value is EMPTY!")
            print("\nIgnoring it.")
            time.sleep(5)
            return
        
        path = the_folder_name

        # Check if the path exists
        if os.path.exists(path):
            print(f"The Folder '{path}' exists.")
        else:
            print(f"The Folder '{path}' does not exist.")
            while True:
              time.sleep(600)
            
            
        #check if 'cookie.txt' exists in this folder
        
        path = the_folder_name + '/cookie.txt'
        
        if os.path.exists(path):
            print(f"The Cookie file 'cookie.txt' exists in the Folder, {the_folder_name}")
        else:
            print(f"The Cookie file 'cookie.txt' DOES NOT exist in the Folder, {the_folder_name}")
            while True:
              time.sleep(600)
              
        
        with open(path, 'r', encoding='utf-8') as f:
                 
            data_JSON = f.read()

            
        #COOKIE DATA
        
        if data_JSON == None or data_JSON == '':
            print("\nThe Cookie Value Is Empty!")
            print("\nExiting This Browser Session.")
            time.sleep(5)
            return
              
              
        #check if 'text.txt' exists in this folder
        
        path = the_folder_name + '/text.txt'
        
        if os.path.exists(path):
            print(f"The file 'text.txt' exists in the Folder, {the_folder_name}")
        else:
            print(f"The file 'text.txt' DOES NOT exist in the Folder, {the_folder_name}")
            while True:
              time.sleep(600)
              
              
        with open(path, 'r', encoding='utf-8') as f:
            
            text_content = f.read()
            
            
        
        
              
        folder_path = the_folder_name

        # List all files in the folder
        files = os.listdir(folder_path)
        print("\nfiles : ",files)
              
        # Check for image files (for example, checking for .jpg, .png, .jpeg extensions)
        image_extensions = ['.jpg', '.jpeg', '.png']  # Add more extensions as needed

        image_files = [
            file for file in files
            if (
                os.path.isfile(os.path.join(folder_path, file))
                and os.path.splitext(file)[1].lower() in (ext.lower() for ext in image_extensions)
            )
        ]

        if len(image_files) > 0:
            print(f"Image files found in the folder, {folder_path}")
            for image_file in image_files:
                image_file = the_folder_name + f'/{image_file}'
                # Convert relative path to absolute path
                image_file_absolute_path = os.path.abspath(image_file)
                print(f"The image file : {image_file}")
                print("\nimage_file_absolute_path : ",image_file_absolute_path)
        else:
            print(f"\nNo image files found in the folder, {folder_path}","\nThe image files must have an extension out of ['.jpg', '.jpeg', '.png']")
            while True:
              time.sleep(600)
        
        
        
    except Exception as e:
        print(e)
        while True:
            time.sleep(600)
        
        
    print("\nOpening Chrome....")    
    
    while True:
        
        try:
            capabilities = DesiredCapabilities.CHROME.copy()
            capabilities["goog:loggingPrefs"] = {"performance": "ALL"}

            #options = uc.ChromeOptions()
            
            chrome_options = Options()
            
            #chrome_options.add_extension("ublock_origin.crx")

            # CHROME DISABLE NOTIFICATION
            
            chrome_options.add_argument('--ignore-ssl-errors=yes')
            chrome_options.add_argument('--ignore-certificate-errors')
            chrome_options.add_argument('--disable-notifications')
            
            
            """driver = uc.Chrome(options=options, service=Service(ChromeDriverManager().install()),
                                desired_capabilities=capabilities, use_subprocess=True)"""
                                
            driver = webdriver.Chrome(options=chrome_options)
            # Set page load timeout: if the page is not loaded within this time limit it will show an TimeoutException error.
            driver.set_page_load_timeout(120)
            wait = WebDriverWait(driver, 10)
            print("\nYour chrome browser version:", driver.capabilities['browserVersion'])
            break
        except Exception as e:
            print("\n",e)
            print("\n\n~~~Your chrome browser might need an update, check if there is an update available.~~~\n\n")
            time.sleep(600)
            exit()
            
            
            
    def disable_password(driver):
        wait = WebDriverWait(driver, 5)
        try:
            driver.get("chrome://settings/passwords?search=password")
            time.sleep(1)
            driver.execute_script(
                '''document.querySelector("body > settings-ui").shadowRoot.querySelector("#main").shadowRoot.querySelector("settings-basic-page").shadowRoot.querySelector("#basicPage > settings-section.expanded > settings-autofill-page").shadowRoot.querySelector("#passwordSection").shadowRoot.querySelector("#passwordToggle").shadowRoot.querySelector("#outerRow").click();''')
            print("\nPassword Save option Disabled")
            time.sleep(2)
        except:
            try:
                driver.get("chrome://password-manager/settings")
                time.sleep(1)
                driver.execute_script(
                    '''document.querySelector("body > password-manager-app").shadowRoot.querySelector("#settings").shadowRoot.querySelector("#passwordToggle").shadowRoot.querySelector("#outerRow").click();''')
                print("\nPassword Save option Disabled")
                time.sleep(2)
            except:
                print("\nPassword Save option Could not be Disabled")







    def turn_on_popup(driver):
        while True:
            try:
                driver.get("chrome://settings/content/popups?search=pop")
                time.sleep(2)
                driver.execute_script("""document.querySelector("body > settings-ui").shadowRoot.querySelector("#main").shadowRoot.querySelector("settings-basic-page").shadowRoot.querySelector("#basicPage > settings-section.expanded > settings-privacy-page").shadowRoot.querySelector("#pages > settings-subpage.iron-selected > settings-category-default-radio-group").shadowRoot.querySelector("#enabledRadioOption").shadowRoot.querySelector("#radioCollapse").click();""")
                print("\nTurned on the popup and redirect setting.")
                time.sleep(2)
                break
            except:
                print("\nFaced some issues while turning on the popup and redirect setting.\n") 
                time.sleep(5)



    disable_password(driver)

    turn_on_popup(driver)     
            



    cookie_login_failed = 0
    
    while True:

        try:

            try:
                print("\nVisiting The Site")
                driver.get('https://www.quora.com/')
                try:
                    checking_if_page_loaded_successfully = driver.find_element(By.XPATH, '//*[@id="email"]')
                    print("\nPage Loaded Successfully.")
                except:
                    login_successfully = WebDriverWait(driver,1).until(EC.presence_of_element_located((By.CSS_SELECTOR, '.base___StyledClickWrapper-lx6eke-1.kXuXkr.puppeteer_test_add_question_button')))
                    print("\nAlready logged in")
                    break

            except:
                print("\nFaced Some Issues. Retrying in 5 seconds")
                time.sleep(5)
                #continue

            print("\nNow Adding Cookies...")

            

            # Convert JSON string to dictionary
            data_dict_list = json.loads(data_JSON)
            #print(type(data_dict_list))
            cookie_list = []

            cookie_dict = {}
            mykeys = ['name', 'value', 'domain', 'path']
            for index, dictionary in enumerate(data_dict_list):
                #print("\n")
                #print(f"{index} : ",dictionary)
                for key, value in dictionary.items():
                    if key == 'name':
                        cookie_dict[key] = f"{value}"
                    if key == 'value':
                        cookie_dict[key] = f"{value}"
                    if key == 'domain':
                        cookie_dict[key] = f"{value}"
                    if key == 'path':
                        cookie_dict[key] = f"{value}"
                sorted_cookie_dict = {i: cookie_dict[i] for i in mykeys}
                cookie = sorted_cookie_dict
                cookie_list.append(sorted_cookie_dict)
            
            print(f"\nAdding Cookie for Row No. : {the_row}")
            
            #print("\ncookie_list : ",cookie_list)
            
            #print(cookie_list)
            try:
                for i in cookie_list:
                    driver.add_cookie(i)
            except Exception as e:
                print(e)
                print("\nCookie couldn't be added properly. Retrying.")
                continue
                
                
            print("\nCookie added")
            
            time.sleep(2)
            
            print("\nRefreshing..")

            driver.get('https://www.quora.com/')
            
            time.sleep(1)
            
            try:
                login_successfully = WebDriverWait(driver,4).until(EC.presence_of_element_located((By.CSS_SELECTOR, '.base___StyledClickWrapper-lx6eke-1.kXuXkr.puppeteer_test_add_question_button')))
                print("\n........")
            except:
                print(f"\nCouldn't login Properly, Row No: {the_row}, Either Cookie issue or Network Issue")
                print("\nRETRYING!")
                time.sleep(1)
                cookie_login_failed = cookie_login_failed + 1
                if cookie_login_failed > 4:
                    while True:
                        print("\nClosing & Opening New Browser Instance.")
                        try:
                            driver.close()
                            driver.quit()
                        except:
                            pass
                        try:
                            print(f"\nRe-Initiate Browser for row_number : {the_row} ......\n")
                            
                            print("\nOpening Chrome....") 
                        
                            #options = uc.ChromeOptions()
                            
                            chrome_options = Options()
                            
                            #chrome_options.add_extension("ublock_origin.crx")

                            # CHROME DISABLE NOTIFICATION

                            chrome_options.add_argument('--ignore-ssl-errors=yes')
                            chrome_options.add_argument('--ignore-certificate-errors')
                            chrome_options.add_argument('--disable-notifications')
                            
                            """driver = uc.Chrome(options=options, service=Service(ChromeDriverManager().install()),
                                                desired_capabilities=capabilities, use_subprocess=True)"""
                                                
                            driver = webdriver.Chrome(options=chrome_options)
                            # Set page load timeout: if the page is not loaded within this time limit it will show an TimeoutException error.
                            driver.set_page_load_timeout(120)
                            wait = WebDriverWait(driver, 10)
                            print("\nYour chrome browser version:", driver.capabilities['browserVersion'])
                            
                            disable_password(driver)
                            
                            turn_on_popup(driver)
                            
                            break
                        
                        except Exception as e:
                            print(f"\nAfter closing the browser, failed To Re-Initiate Browser for row_number : {the_row}. Retrying!")
                    cookie_login_failed = 0
                continue
            
            print("\nPAGE logged In Successfully.")

            break

        except Exception as e:
            print(e)
            time.sleep(2)
            continue
        
        
    ########################
    while True:
        try:
            multi_window = driver.window_handles

            number_multi_window = len(multi_window)

            print("Opened number of window(s): ", number_multi_window)

            while number_multi_window > 1:
                driver.switch_to.window(driver.window_handles[-1])
                time.sleep(1)
                print("Closing the window")
                driver.close()
                print("The window is closed")
                multi_window = driver.window_handles
                number_multi_window = len(multi_window)
                print("The number of window(s): ", number_multi_window)
                time.sleep(1)
                
            print("\nOpening a second tab for google translate")
            # Open a new window
            #driver.execute_script("window.open('');")
            #driver.execute_script("window.open('', 'new_window')")
            driver.execute_script("window.open('', '_blank')")
            time.sleep(1)
            driver.switch_to.window(driver.window_handles[1])
            time.sleep(1)
            driver.get('https://translate.google.com/')
            time.sleep(2)
            driver.refresh()
            time.sleep(2)
            
            break
        except Exception as e:
            print(e)
    
    #########################
    
    for language in all_languages:
        
        print("\nCurrent_language : ", language)
        
    
    
        while True:
            
            try:
                
                while True:
                    
                    try:
                
                        print("\nVisiting Quora Home Page")
                        
                        driver.switch_to.window(driver.window_handles[0])
                        
                        language_name = language.strip()

                        # Search for the language by name and retrieve the ISO 639-1 code
                        language_code = next((lang.alpha2 for lang in languages if lang.name.lower() == language_name.lower()), None)

                        print(f"Language code for '{language_name}': {language_code}")
                        
                        if language_code == None:
                            print("\nNo Language match found to generate language_code")
                            raise Exception
                        
                        
                        language_code = str(language_code).strip().lower()
                        
                        quora_home_page_url = "https://www.quora.com/".replace('www', f'{language_code}')
                        
                        print("\nquora_home_page_url : ",quora_home_page_url)
                        
                        driver.get(quora_home_page_url)
                        
                            
                        try:
                            
                            print("\nLooking for the 'Add' button for the targeted language : ", language_name)
                            
                            add_targeted_btns =  wait.until(
                                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".puppeteer_test_button_text")))
                            
                            print("\nlen(add_targeted_btns) : ", len(add_targeted_btns))
                            
                            if len(add_targeted_btns) == 1:
                            
                                add_targeted_btn = add_targeted_btns[0]
                                
                                try:
                                    print("\nClicking")
                                    add_targeted_btn.click()
                                    print("\nClicked.")

                                except:
                                    
                                    driver.execute_script("arguments[0].click();", add_targeted_btn)
                                    print("\nClicked..")
                                    
                                time.sleep(3)
                            
                            else:
                                print("\nNo need to confirm the language.")
                                
                        except Exception as e:
                            pass
                        
                        
                        #####   translate #############
                        while True:
                            textarea_filed_found = False
                            try:
                                print("\nSwitching window to Google translate.")
                                driver.switch_to.window(driver.window_handles[1])
                                time.sleep(1)
                                print("\nlanguage_name : ",language_name)
                                
                                # Example usage:
                                selected_language = language_name
                                selected_language_google_translate_code = get_language_code(selected_language)
                                if selected_language_google_translate_code == None:
                                    selected_language_google_translate_code = 'en'
                                print(f"The code for {selected_language} is {selected_language_google_translate_code}")
                                
                                
                                google_translate_language_selected_url = f'https://translate.google.com/?sl=auto&tl={selected_language_google_translate_code}&op=translate'
                                driver.get(google_translate_language_selected_url)
                                print("\ngoogle_translate_language_selected_url : ",google_translate_language_selected_url)
                                
                                print("\nLooking for 'textarea'")
                                textarea_fields = wait.until(
                                    EC.presence_of_all_elements_located((By.TAG_NAME, "textarea")))
                                print("\nlen(textarea_fields) : ", len(textarea_fields))
                                for index, textarea_field in enumerate(textarea_fields):
                                    print("\nindex : ",index)
                                    try:
                                        the_attribute_value = textarea_field.get_attribute('aria-label')
                                        print("\naria-label, the_attribute_value : ", the_attribute_value)
                                        if 'Source text' in the_attribute_value:
                                            print("\nFound the textarea")
                                            textarea_field = textarea_fields[index]
                                            textarea_filed_found = True
                                            break 
                                    except:
                                        pass
                                    
                                if textarea_filed_found == True:
                                    
                                    print("\nClicking")
                                    textarea_field.click()
                                    print("\nClicked.")
                                    print("\nClearing the text filed")
                                    time.sleep(1)
                                    action_chain = ActionChains(driver)
                                    action_chain.key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
                                    action_chain.key_down(Keys.CONTROL).send_keys('x').key_up(Keys.CONTROL).perform()
                                    print("\nCleared.")
                                    time.sleep(1)
                                    print("\nEntering text content")
                                    text_content = f"""{text_content}"""
                                    print("\ntext_content : ",text_content)
                                    time.sleep(2)
                                    #pyperclip.copy(text_content)
                                    #clipboard.value = text_content
                                    #action_chain.key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
                                    textarea_field.send_keys(text_content)
                                    print("\nEntered")
                                    time.sleep(5)
                                    print("\nLooking for a button with aria-label attribute of value : 'Copy translation'")
                                    all_buttons = wait.until(
                                    EC.presence_of_all_elements_located((By.TAG_NAME, "button")))
                                    
                                    print("\nlen(all_buttons) : ",len(all_buttons))
                                    
                                    the_copy_result_button_found = False
                                    
                                    for index, the_button in enumerate(all_buttons):
                                        
                                        try:
                                            print("\nindex : ",index)
                                            the_attribute = the_button.get_attribute('aria-label')
                                            print("\nthe_attribute : ",the_attribute)
                                            if 'Copy translation' in the_attribute:
                                                print("\nFound")
                                                the_copy_result_button_found = True
                                                the_copy_result_button = all_buttons[index]
                                                break
                                        except:
                                            pass
                                        
                                        
                                    if the_copy_result_button_found == True:
                                        
                                        """try:
                                            the_copy_result_button.click()
                                            print("\nClicked.")
                                        except:
                                            
                                            driver.execute_script("arguments[0].click();", the_copy_result_button)
                                            print("\nClicked..")
                                            
                                        time.sleep(1)
                                            
                                        #the_copied_result = pyperclip.paste()
                                        the_copied_result = clipboard.value
                                        print("\nthe_copied_result : ",the_copied_result)"""
                                        
                                        js_code_translation = """
                                        
                                        const elements = document.querySelectorAll('.ryNqvb');

                                        let combinedText = '';

                                        elements.forEach(element => {
                                            combinedText += element.textContent + ' ';
                                        });
                                        
                                        return combinedText
                                        
                                        """
                                        
                                        the_copied_result = driver.execute_script(js_code_translation)
                                        
                                        print("\nThe translated text is extracted")
                                        
                                        
                                        break
                                    
                                else:
                                    print("\ntextarea Not Found!")
                                
                            except Exception as e:
                                print(e)
                                time.sleep(5)
                                continue
                                
                                        
                                    

                        print("\nSwitching to the first window")
                        driver.switch_to.window(driver.window_handles[0])
                        time.sleep(1)  
                        
                        print("\nLooking for 'Add question' Button.")
                        
                        add_question_btns = wait.until(
                                        EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".puppeteer_test_add_question_button")))
                        
                        print("\nlen(add_question_btns) : ",len(add_question_btns))
                        
                        print("\nClicking 'Add question' Button.")
                        
                        driver.execute_script("""document.querySelector(".puppeteer_test_add_question_button").click();""")
                        
                        print("\nClicked.")
                        
                        print("\nLooking for 'Create Post' button")
                        
                        all_the_buttons = wait.until(
                                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".qu-dynamicFontSize--button")))
                        
                        print("\nlen(all_the_buttons) : ",len(all_the_buttons))
                        
                        create_post_button_found = False
                        create_post_button_clicked = False
                        
                        if len(all_the_buttons) == 2:
                                    
                            create_post_button_found = True
                            
                            
                            print("\nFound.")
                            
                            print("\nNow Selecting the Button")
                            
                            try:
                                
                                the_button = all_the_buttons[-1].click()
                                
                                print("\nClicked.")
                                create_post_button_clicked = True
                                
                            except Exception as e:
                                print(e)
                                try:
                                    driver.execute_script(f"""document.querySelectorAll(".qu-dynamicFontSize--button")[1].click();""")
                                    print("\nClicked..")
                                    create_post_button_clicked = True
                                except Exception as e:
                                    print(e)
                                    print("\nFAILED TO CLICK !")
                                    create_post_button_clicked = False

                    
                    
                        if create_post_button_found == True and create_post_button_clicked == True :
                            break
                    
                    except Exception as e:
                        print(e)
                        continue
                    
                ############################
                
                try:
                    
                    print("\nDeleting all the existing text & images in the text field")
                    
                    print("\nFinding content_el")
                    
                    content_els = wait.until(
                                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".section")))
                    
                    
                    print("\nlen(content_els) : ", len(content_els))
                    try:
                        content_el = content_els[-1]
                    except Exception as e:
                        print(e)
                        
                        content_els = driver.execute_script("""return document.querySelectorAll('.section')""")
                        print("\nlen(content_els) :- ", len(content_els))
                        content_el= content_els[len(content_els)-1]
                    
                    print("\nFound")
                    
                    try:
                    
                        content_el.click()
                        print("\nClicked.")
                    except:
                        
                        driver.execute_script("arguments[0].click();", content_el)
                        print("\nClicked..")
                    
                    
                    action_chain = ActionChains(driver)
                    action_chain.key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
                    action_chain.key_down(Keys.CONTROL).send_keys('x').key_up(Keys.CONTROL).perform()
                    time.sleep(1)
                    action_chain.key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
                    action_chain.key_down(Keys.CONTROL).send_keys('x').key_up(Keys.CONTROL).perform()
                    time.sleep(1)
                    
                    print("\nCleared all the text & images")    
                    
                    time.sleep(2)
                    
                    
                    file_uploaded = False
                    
                    print("\nFinding the image input Button")
                    
                    file_upload_btn = wait.until(
                                    EC.presence_of_all_elements_located((By.XPATH, "//*[@id='modal_footer_portal_container']/div/div/div/div[1]/div[1]/div[2]")))[0]
                    
                    # The above code is printing the string "Button Found" to the console.
                    print("\nButton Found.")

                    # Execute the JavaScript to inject the file input element
                    js_code = """
                    var fileInput = document.createElement('input');
                    fileInput.type = 'file';
                    fileInput.accept = 'image/*';
                    fileInput.addEventListener('change', function(event) {
                        var selectedFile = event.target.files[0];
                        console.log('Selected file:', selectedFile);
                    });
                    document.body.appendChild(fileInput);
                    """
                    driver.execute_script(js_code)
                    
                    time.sleep(1)

                    # Locate the file input using Selenium
                    file_input = driver.find_element(By.CSS_SELECTOR,'input[type="file"][accept="image/*"]')

                    # Perform actions on the file input, such as sending keys (file path)
                    file_input.send_keys(image_file_absolute_path)
                    
                    print("\nImage Uploaded")
                    
                    time.sleep(2)
                    
                    # Perform actions on the file input, such as sending keys (file path)
                    file_input.send_keys(image_file_absolute_path)
                    
                    print("\nImage Uploaded Again")
                    
                    time.sleep(2)
                    
                    print("\nFinding content_el again")
                    
                    content_el = driver.find_elements(By.CSS_SELECTOR,'.section')[-1]
                    
                    print("\nFound")
                    
                    try:
                    
                       content_el.click()
                       print("\nClicked.")
                    except:
                        
                        driver.execute_script("arguments[0].click();", content_el)
                        print("\nClicked..")
                    
                    
                    print("\nWriting...")
                    
                    the_copied_result = the_copied_result+'\n\n'
                    
                    driver.execute_script(f"""document.querySelectorAll('.section')[document.querySelectorAll('.section').length-1].textContent=`{the_copied_result}`;""")
                    
                    
                    #content_el.send_keys(the_copied_result)
                    
                    time.sleep(2)
                    
                    print("\nNow Adding the link")
                    

                    the_link_to_be_pasted = the_link_to_add
                    
                    pyperclip.copy(the_link_to_be_pasted)
                    
                    action_chain.key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
                    
                    print("\nAdded the link")
                    
                    time.sleep(5)
                    
                    print("\nDeleting the last photo")
                    
                    delete_els = driver.find_elements(By.CSS_SELECTOR, '.section_remove_btn')
                    
                    print("\nlen(delete_els) : ", len(delete_els))
                    
                    delete_el = delete_els[-1]
                    
                    print("\nFound.")
                    
                    try:
                    
                        delete_el.click()
                        print("\nClicked.")
                    except:
                        
                        driver.execute_script("arguments[0].click();", delete_el)
                        print("\nClicked..")
                        
                        
                    time.sleep(2)
                    
                    
                    print("\nFinding Post Button")
                    
                    post_btn_els = wait.until(
                                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".puppeteer_test_modal_submit")))
                    
                    print("\nlen(post_btn_els) : ", len(post_btn_els))
                    
                    
                    post_btn_el = post_btn_els[0]
                    
                    print("\nPost Button Found")
                    
                    """try:
                        post_btn_el.click()
                        print("\nClicked.")
                    except:
                        driver.execute_script("arguments[0].click();", post_btn_el)
                        print("\nClicked..")"""
                        
                        
                    driver.execute_script("""document.querySelector('.puppeteer_test_modal_submit').click();""")
                        
                    

                    time.sleep(5)
                    print("\nposted")
                    
                    print("\nClosing 2nd tab which is the page opened after posting the content.")
                    
                    while True:
                        try:
                    
                            multi_window = driver.window_handles

                            number_multi_window = len(multi_window)

                            print("Opened number of window(s): ", number_multi_window)
                            k = 0
                            while number_multi_window > 2:
                                k = k+1
                                print("\nk : ",k)
                                driver.switch_to.window(driver.window_handles[0])
                                time.sleep(1)
                                driver.switch_to.window(driver.window_handles[k])
                                time.sleep(1)
                                if 'Translate' not in driver.title:
                                    
                                    print("Closing the window")
                                    driver.close()
                                    print("The window is closed")
                                    multi_window = driver.window_handles
                                    number_multi_window = len(multi_window)
                                    print("The number of window(s): ", number_multi_window)
                                
                            driver.switch_to.window(driver.window_handles[0])
                            break
                            
                        except Exception as e:
                            print(e)
                            continue
                    

                    
                
                except Exception as e:
                    print("\nError...")
                    print(e)
                    print("\nWaiting for 4 seconds before retrying.") 
                    time.sleep(5)
                    continue
                
                ############################
                
                print("\nChecking if it is posted or in draft mode")
                
                try:
                    the_url_here = quora_home_page_url + '/drafts'
                    
                    print("\nthe_url_here : ",the_url_here)
                    
                    print("\nvisiting....")
                    
                    driver.get(the_url_here)
                    
                    time.sleep(2)
                    
                    print("\nLooking for 'Discard' Btn")
                    
                    edit_btn_found = False
                    
                    discard_btns =  wait.until(
                                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".q-click-wrapper")))
                    
                    print("\nlen(discard_btns) : ", len(discard_btns))
                    
                    
                    driver.execute_script("""document.querySelectorAll('.q-click-wrapper')[document.querySelectorAll('.q-click-wrapper').length-2].click();""")
                    
                    edit_btn_found = True
                    
                    print("\nFound and Clicked.")
                    
                    time.sleep(2)
                    
                    print("\nlooking for Discard popup Button")
                    
                    discard_popup_btns =  wait.until(
                                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".q-click-wrapper")))
                    
                    discard_popup_btns_len = len(discard_popup_btns)
                    
                    print("\ndiscard_popup_btns_len : ",discard_popup_btns_len)
                    
                    driver.execute_script("""document.querySelectorAll('.q-click-wrapper')[document.querySelectorAll('.q-click-wrapper').length-1].click();""")
                    
                    print("\nDiscard popup Button clicked.")
                    
                    
                except Exception as e:
                    print(e)
                    
                time.sleep(5)
                
                print("\nGoing to the next language.")
                break
                
                
            except Exception as e:
                print(e)
                print("\nWaiting for 5 seconds before retrying.") 
                ws_Settings.cell(row=the_row, column=3, value=str(e))
                wb_Settings.save('Setting.xlsx')
                time.sleep(5)
                print("\nWait is over.")

    ws_Settings.cell(row=the_row, column=3, value='Completed')
    wb_Settings.save('Setting.xlsx')
    time.sleep(1)








################################################################################################################################           





#########################################################

if __name__ == '__main__':
    
    multiprocessing.freeze_support()  
    

    print("\nChecking if the 'Setting.xlsx' file is present in the Current Directory.")

    row_Settings_max = 0

    try:
        wb_Settings = openpyxl.load_workbook("Setting.xlsx")
        ws_Settings = wb_Settings['Sheet1']
        row_Settings_max = ws_Settings.max_row
        
        print("\nNumber of the rows in the excel file, 'Setting.xlsx' : ", row_Settings_max)
        
        print("\nNumber of Accounts provided in the excel file, 'Setting.xlsx' : ",row_Settings_max-1)

    except Exception as e:
        print(e)
        print("\nThe 'Setting.xlsx' is NOT present in the Current Directory")
        time.sleep(600)
        exit()
        


    #########################################################

    if row_Settings_max < 2:
        print("\nNumber of total rows is less than 2.\nIt must be more or equal to 2 as the data-reading starts from row no.2 ")
        time.sleep(600)
        exit()



    print("\nLooking for 'the_link_to_add.txt' in the main folder")

    try:

        with open('the_link_to_add.txt', 'r', encoding='utf-8') as f:
            
                the_link_to_add = f.readline()
                
                print("\nthe_link_to_add : ",the_link_to_add)
            
            
    except Exception as e:
        print("\nThe file, 'the_link_to_add.txt' doesn't exist!")
        print(e)
        while True:
            time.sleep(600)
            
            
            
            
    print("\nLooking for 'languages.txt' in the main folder")

    try:

        with open('languages.txt', 'r', encoding='utf-8') as f:
            
                all_languages = [language.strip() for language in f.readlines()]
                
                print("\nall_languages : ",all_languages)
                
            
            
    except Exception as e:
        print("\nThe file, 'languages.txt' doesn't exist!")
        print(e)
        while True:
            time.sleep(600)
    
    
    
    
    while True:

        start_from = input("\nEnter the row no. you want to start from. (minimum=2)\n: ").strip()
        
        print("\nYou have entered : ", start_from)
        
        try:
            
            start_from = int(start_from)
            
            if start_from <= 1:
                raise Exception
            
        except:
            print("\nplease enter an Integer which is greater than 1")
            continue
        
        
        if start_from > row_Settings_max:
            print(f"\nEntered row no. : {start_from} is greater than the maximum row no. : {row_Settings_max} in the excel file.")
            print(f"\nEnter an Integer less than or equal to the maximum row no. : {row_Settings_max}")
            continue
        
        
        print(f"\n You have selected row no. : ",start_from)
        break
        
        
    #########################################################


    while True:

        end_at = input(f"\nEnter the row no. you want to END at. (maximum={row_Settings_max})\n: ").strip()
        
        print("\nYou have entered : ", end_at)
        
        try:
            
            end_at = int(end_at)
            
            if start_from > end_at:
                raise Exception
            
        except:
            print(f"\nplease enter an Integer greater than or equal to {start_from}")
            continue
        
        
        if end_at > row_Settings_max:
            print(f"\nEntered row no. : {end_at} is greater than the maximum row no. : {row_Settings_max} in the excel file.")
            print(f"\nEnter an Integer less than or equal to the maximum row no. : {row_Settings_max}")
            continue
        
        
        print(f"\n You have selected row no. : ",end_at)
        break




    #########################################################
    
    
    
    
    
    #########################################################




    print("\nNow checking if Each Folder consists : 'cookie.txt', 'text.txt', an image ")

    the_row_list = []

    for the_row in range(start_from, end_at+1):
        
        path = ''
        the_folder_name = ''
        
        try:
            print("\nThe Row No. : ", the_row)
            
            the_row_list.append(the_row)
            
            the_folder_name = ws_Settings.cell(row= the_row,column= 2).value
            
            print("\nthe_folder_name : ",the_folder_name)
            
            if the_folder_name == '' or the_folder_name == None :
                print("\nThe folder name's value is EMPTY!")
                print("\nIgnoring it.")
                time.sleep(1)
                continue
            
            path = the_folder_name

            # Check if the path exists
            if os.path.exists(path):
                print(f"The Folder '{path}' exists.")
            else:
                print(f"The Folder '{path}' does not exist.")
                while True:
                  time.sleep(600)
                
                
            #check if 'cookie.txt' exists in this folder
            
            path = the_folder_name + '/cookie.txt'
            
            if os.path.exists(path):
                print(f"The Cookie file 'cookie.txt' exists in the Folder, {the_folder_name}")
            else:
                print(f"The Cookie file 'cookie.txt' DOES NOT exist in the Folder, {the_folder_name}")
                while True:
                  time.sleep(600)
                
                
            #check if 'text.txt' exists in this folder
            
            path = the_folder_name + '/text.txt'
            
            if os.path.exists(path):
                print(f"The file 'text.txt' exists in the Folder, {the_folder_name}")
            else:
                print(f"The file 'text.txt' DOES NOT exist in the Folder, {the_folder_name}")
                while True:
                  time.sleep(600)
                
            folder_path = the_folder_name

            # List all files in the folder
            files = os.listdir(folder_path)     
                
            # Check for image files (for example, checking for .jpg, .png, .jpeg extensions)
            image_extensions = ['.jpg', '.jpeg', '.png']  # Add more extensions as needed

            image_files = [
                file for file in files
                if (
                    os.path.isfile(os.path.join(folder_path, file))
                    and os.path.splitext(file)[1].lower() in (ext.lower() for ext in image_extensions)
                )
            ]

            if len(image_files) > 0:
                print(f"Image files found in the folder, {folder_path}")
                for image_file in image_files:
                    print(f"The image file : {image_file}")
            else:
                print(f"\nNo image files found in the folder, {folder_path}","\nThe image files must have an extension out of ['.jpg', '.jpeg', '.png']")
                while True:
                  time.sleep(600)
            
            
            
        except Exception as e:
            print(e)
            while True:
                time.sleep(600)
    
    
    print("\nthe_row_list : ",the_row_list) 

    time.sleep(2)
 
 
    ###############################################################
    
    # Create a manager to manage the clipboard
    manager = Manager()
    clipboard = manager.Value(str, "")  # Creating a shared string variable
    
    ###############################################################
    
            
    with multiprocessing.Pool(processes=len(the_row_list)) as pool:
        partial_main_program = functools.partial(quora_post, the_link_to_add=the_link_to_add,all_languages=all_languages, ws_Settings=ws_Settings, wb_Settings=wb_Settings,clipboard=clipboard)
        pool.map(partial_main_program, the_row_list)
        
    print("\nThe Task is Completed!!!")    
    while True:
        time.sleep(600)











           
           
            
            

